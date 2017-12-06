"""
.py
Get the stories and defects in an iteration
Put them in a csv file
Print summaries to the screen
------------------------------------------------
001 Can print stories and defects to the screen
002 Restructured to use main() and put data into a dictionary
003 Responses are dictionaries, but all data in a list
004 Restructured with smaller routines & ability to read prior 
        results from a file to speed development
005 Added new calc for zero pt and moved % done calcs into the dictionary
        Changed read file to be the same as the output csv
006 Command line switches started, reformat to 4 width, not 2
007 more command line switches
008 Excel output
009 Add iteration command line
010 Automagic iteration creation

Needed
error handling with results is none
overwrite or new name on output
environment variables: project/repository, iteration, path to output
?query file
multiple iterations in one workbook
?formatted workbook
change project on arg
simplified iteration names
multiple iterations in one file
control screen output
separate routine for command line processing
"""
from __future__ import print_function
import os
import sys
import csv
import getopt
import datetime
from pyral import Rally, rallySettings, rallyWorkset
from openpyxl import Workbook


gblResults = [] # data results
gRally = object
#gblIteration = 'Iteration.Name = "S03#2017-02-01/02-14"'
gblIteration = ""
gblReport = {}


def getCurrentIteration():
    iter1start = datetime.datetime(2017, 1, 4)
    today = datetime.datetime.today()
    inIteration = int((today - iter1start).days / 14.) + 1 
    return getIterationLabel(inIteration)

def getPreviousIteration():
    iter1start = datetime.datetime(2017, 1, 4)
    today = datetime.datetime.today()
    inIteration = int((today - iter1start).days / 14.) 
    return getIterationLabel(inIteration)

def getIterationLabel(inIteration):
    iter1start = datetime.datetime(2017, 1, 4)
    iter1end = datetime.datetime(2017, 1, 17)
    today = datetime.datetime.today()
    deltaDays = datetime.timedelta(days=((inIteration-1)*14))
    iterstart = iter1start + deltaDays
    iterend = iter1end + deltaDays
    strIterLabel = ("S{:02d}#2017-{:02d}-{:02d}/{:02d}-{:02d}").format(
            inIteration, iterstart.month, iterstart.day, iterend.month, iterend.day)
    return strIterLabel
    
    

def printHelp():
    print("Command line options")
    print("    -f                  Use the output file out.csv, no Rally refresh")
    print("    -h | -help          Print help")
    print("    -i | -input file    Use the named input file; turns off refresh and default is out.csv")
    print("    -o | -output file   The output file; default out.csv")
    print("    -p path             The output path; default current directory")
    print("    -r true/false       Refresh the data t/f; default refesh from Rally")
    print("    -s true/false       Refresh the data t/f; default refesh from Rally")



def setupRally():
    global gRally
    server, user, password, apikey, workspace, project = 'rally1.rallydev.com', 'christopher.eberhardt@sabre.com', 'Unif0rm.21', '', 'Sabre Production Workspace', 'GetThere' 
    project = 'Traveler Experience' 
    gRally = Rally(server, user, password, workspace=workspace, project=project)

def getgResultsQuery(queryList):
    global gRally, gblResults
    for queryItem in queryList:
        #response = gRally.get(queryItem, query=gblIteration) 
        response = gRally.get(queryItem, fetch=True, projectScopeDown=True, query=gblIteration) 

        for item in response:
            if hasattr(item, 'Owner') and hasattr(item.Owner, 'Name'):
                owner = item.Owner.Name
            else:
                owner = ""
            if hasattr(item, 'TeamFeature') and hasattr(item.TeamFeature, 'Name'):
                teamfeature = item.TeamFeature.Name
            else:
                teamfeature = ""

            if hasattr(item, 'Iteration') and hasattr(item.Iteration, 'Name'):
                iteration = item.Iteration.Name
            else:
                iteration = "What?"

            gblResults.append( {
                        'FormattedID': item.FormattedID, 
                        'Name': item.Name, 
                        'PlanEstimate': item.PlanEstimate, 
                        'TeamFeature': teamfeature, 
                        'ScheduleState': item.ScheduleState, 
                        'ProjectName': item.Project.Name, 
                        'IterationName': item.Iteration.Name, 
                        'Owner': owner } )


def getgResultsFile(filename):
    global gRally, gblResults
    with open(filename) as f:
        gblResults = [{k: v for k, v in row.items()}
            for row in csv.DictReader(f, skipinitialspace=True)]

def writeCSV(filename):
    global gRally, gblResults
    with open(filename, 'wb') as csvfile:
        fieldnames = ['FormattedID', 'Name', 'PlanEstimate', 'TeamFeature',
                        'ScheduleState', 'ProjectName', 'IterationName', 'Owner']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, dialect='excel')
        writer.writeheader()
        for story in gblResults:
                writer.writerow(story)

def writeExcel():
    global gblResults, gblReport

    wb = Workbook()

    today = datetime.datetime.today()
    strDate = ("{:4d}.{:02d}.{:02d}").format(today.year, today.month, today.day)
    ws = wb.create_sheet(strDate)

    fieldnames = ['FormattedID', 'Name', 'PlanEstimate', 'TeamFeature',
                    'ScheduleState', 'ProjectName', 'IterationName', 'Owner']
    row = 1
    col = 1
    for field in fieldnames:
        ws.cell(column=col, row=row, value=field)
        col += 1
    row += 1

    for story in gblResults:
        col = 1
        #ws.cell(column=col, row=row, value=story)
        #col += 1
        for field in fieldnames:
            ws.cell(column=col, row=row, value=story[field])
            col += 1
        row += 1

    # write the summary xx
    ws = wb.create_sheet("Summary")
    row = 1
    orderAttrib = [ 'WIP', 'StoryCount', 'StoryDone', 'Points', \
            'PointsDone', 'DefectCount', 'DefectDone', \
            'StoryPercent', 'PointsPercent', 'ZeroPoints', \
            'StoryZero', 'DefectZero']
    teams = gblReport.keys()
    teams.sort()
    col = 2
    for team in teams:
        #ws.cell(column=col, row=row, value=team[9:11])
        ws.cell(column=col, row=row, value=team)
        col += 1
    row += 1
    for attrib in orderAttrib:
        col = 1
        ws.cell(column=col, row=row, value=attrib)
        for team in teams:
            col += 1
            ws.cell(column=col, row=row, value=gblReport[team][attrib])
        row += 1
    ws.column_dimensions["A"].width = 15


    ws = wb.get_sheet_by_name("Sheet")
    if ws is not None:
        wb.remove_sheet(ws)

    fname = "PyRally." + strDate + ".xlsx"
    wb.save(fname)

def processRecord(record):
    global gblReport
    projectName = record['ProjectName']
    if record['FormattedID'][0] == "U":
        gblReport[projectName]['StoryCount'] += 1
    else:
        gblReport[projectName]['DefectCount'] += 1

    if record['PlanEstimate']:
        gblReport[projectName]['Points'] += float(record['PlanEstimate'])
        if float(record['PlanEstimate']) == 0.0:
            gblReport[projectName]['ZeroPoints'] += 1
            if record['FormattedID'][0] == "U":
                gblReport[projectName]['StoryZero'] += 1
            else:
                gblReport[projectName]['DefectZero'] += 1

    else:
        gblReport[projectName]['ZeroPoints'] += 1
        if record['FormattedID'][0] == "U":
            gblReport[projectName]['StoryZero'] += 1
        else:
            gblReport[projectName]['DefectZero'] += 1

    if record['ScheduleState'] == "Completed" or \
         record['ScheduleState'] == "Accepted":
        if record['FormattedID'][0] == "U":
            gblReport[projectName]['StoryDone'] += 1
        else:
            gblReport[projectName]['DefectDone'] += 1
        if record['PlanEstimate']:
            gblReport[projectName]['PointsDone'] += float(record['PlanEstimate'])
    elif record['ScheduleState'] == "In-Progress":
            gblReport[projectName]['WIP'] += 1

    if gblReport[projectName]['Points'] > 0:
        x = gblReport[projectName]['PointsDone'] * 100.0 / gblReport[projectName]['Points'] 
        x = int(x*10) / 10.0 # put in one decimal point percentage
    else: x = 0.0
    gblReport[projectName]['PointsPercent'] = x

    if gblReport[projectName]['StoryCount'] > 0:
        x = gblReport[projectName]['StoryDone'] * 100.0 / gblReport[projectName]['StoryCount'] 
        x = int(x*10) / 10.0 # put in one decimal point percentage
    else: x = 0.0
    gblReport[projectName]['StoryPercent'] = x

def analyzeIteration():
    global gRally, gblResults
    global gblReport
    for record in gblResults:
        projectName = record['ProjectName']
        if projectName not in gblReport:
            gblReport[projectName] = {}
            gblReport[projectName]['StoryCount'] = 0
            gblReport[projectName]['Points'] = 0.0
            gblReport[projectName]['StoryDone'] = 0
            gblReport[projectName]['PointsDone'] = 0.0
            gblReport[projectName]['DefectCount'] = 0
            gblReport[projectName]['DefectDone'] = 0
            gblReport[projectName]['WIP'] = 0
            gblReport[projectName]['StoryPercent'] = 0.0
            gblReport[projectName]['PointsPercent'] = 0.0
            gblReport[projectName]['ZeroPoints'] = 0
            gblReport[projectName]['StoryZero'] = 0
            gblReport[projectName]['DefectZero'] = 0
        processRecord(record)

    
def printSummaryReport():
    print( "There are ", len(gblResults), " stories and defects")
    orderAttrib = [ 'WIP', 'StoryCount', 'StoryDone', 'Points', \
            'PointsDone', 'DefectCount', 'DefectDone', \
            'StoryPercent', 'PointsPercent', 'ZeroPoints', \
            'StoryZero', 'DefectZero']
    teams = gblReport.keys()
    teams.sort()
    #print ("....+"*14)
    print ("{:>15s}".format(""), end="")
    for team in teams:
        print ("{:>6s}".format(team[9:11]), end="")
    print ()
    for attrib in orderAttrib:
        print("{:>15s}".format(attrib), end="")
        for team in teams:
            print("{:>6s}".format(str(gblReport[team][attrib])), end="")
        print ()


def main():
    global gblResults, gRally, gblIteration, gblReport

    bRefresh = True # Get from Rally, -f to get from out.csv
    outFile = 'out.csv'
    inFile = 'out.csv'
    outPath = "."

    #### Argument processing
    try:
        opts, args = getopt.getopt(sys.argv[1:], "fhi:o:p:r:s:", 
                ["help", "input=", "output=", "sprint="])
    except getopt.GetoptError as err:
        # print help information and exit:
        print (str(err))  # will print something like "option -a not recognized"
        printHelp()
        sys.exit(2)
    output = None
    for o, a in opts:
        if o == "-f":
            bRefresh = False

        elif o in ("-h", "--help"):
            printHelp()
            sys.exit()

        elif o in ("-i", "--input"):
            inFile = a

        elif o in ("-o", "--output"):
            outFile = a

        elif o == "-p":
            outPath = a

        elif o == "-r":
            if a[0] in ("T", "t"):
                bRefresh = True
            else:
                bRefresh = False

        elif o in ("-s", "--sprint"):
            gblIteration = 'Iteration.Name = "' + a + '"'

        else:
            assert False, "unhandled option"


    #print (os.environ['ITER'])
    if bRefresh:
        if gblIteration == "":
            print ("Error No iteration specified")
            sys.exit(2)

        print( "Logging in...")
        setupRally()

        print( "Querying...")
        QueryList = ['UserStory', 'Defect']
        getgResultsQuery(QueryList)

    else:
        getgResultsFile(inFile)
        

    analyzeIteration()
    printSummaryReport()

    # write the results
    #print ("Writing to " + str(outPath + "/" + outFile))
    writeCSV(outPath + "/" + outFile)
    writeExcel()


if __name__ == '__main__':
    gblIteration = 'Iteration.Name = "' + getCurrentIteration() + '"'
    execfile("config.py")
    main()
