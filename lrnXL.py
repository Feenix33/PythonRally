"""
lrnXL Excel experiments tied to Test1.xlsx
01 - Read the sheet and convert to list
02 - See if we can change the value
03 - Iterate through ws and convert to list
04 - Sum a column with filtering
05 - Sum a column by team names using dictionary
06 - Summarize an iteration output
07 - Add Summary sheet to existing XL (at least 27 times ;-)
08 - Write summary (Combine 06 and 07) but using helper functions
09 - Append original data set and keep 08 experiment

Renamed this file to ppIteration
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import pprint
import datetime

#fname = "Test.xlsx"
#fname = "Test1.xlsx"
#fname = "Iter.xlsx"
fname = "IterA.xlsx"

gAttribs = ['Stories.Total',
        'Stories.Zero',
        'Stories.Done',
        'Stories.NotDone',
        'Stories.Percent',
        'Defects.Total',
        'Defects.Done',
        'Defects.NotDone',
        'Points.Total',
        'Points.Done',
        'Points.NotDone',
        'Points.Percent']
############
#gFieldnames = ['FormattedID', 'Name', 'PlanEstimate', 'TeamFeature.FormattedID', 
#        'TeamFeature.Name', 'ScheduleState', 'Project.Name', 'Iteration.Name', 
#        'Owner.Name', 'CreationDate']
#gDoneStates = ['Accepted','Completed','Released-to-Production']
nodata = []
noanly = {}

def addExtendedWS(ws, inDat):
    #use exception handling to punch out if already done
    try:
        noNeed = inDat[0].index('Type')
        return
    except:
        pass
    colID = inDat[0].index('FormattedID')
    colState = inDat[0].index('ScheduleState')
    colIter = inDat[0].index('Iteration.Name')

    outType = ws.max_column + 1
    outDone = outType + 1
    outIter = outDone + 1
    atRow = 2
    ws.cell(row=1, column=outType).value = "Type"
    ws.cell(row=1, column=outDone).value = "Done"
    ws.cell(row=1, column=outIter).value = "Iteration.Sortable"

    for row in inDat[1:]:
        ws.cell(row=atRow, column=outType).value = row[colID][0:2]
        ws.cell(row=atRow, column=outDone).value = isDone(row[colState])
        try:
            ws.cell(row=atRow, column=outIter).value = row[colIter][4:8]+"#"+row[colIter][0:3]
        except TypeError:
            pass
        atRow += 1

def addTypeCol(inDat):
    # add user story or defect to passed data
    # assumes the data has equal length rows
    colID = inDat[0].index('FormattedID')
    inDat[0].append("Type")
    for row in inDat[1:]:
        row.append(row[colID][0:1])

def isDone(state):
    lDoneStates = ['Accepted','Completed','Released-to-Production']
    return state in lDoneStates

def readWSData(ws):
    ldata = []
    for row in ws.values:
        ldata.append(list(row))
    return ldata

def analyzeData(pastDat):
    colTeam = pastDat[0].index("Project.Name")
    teamNames = list(set(row[colTeam] for row in pastDat[1:]))

    empty = dict.fromkeys(gAttribs, 0)
    lanly = dict.fromkeys(teamNames, None)
    for key in lanly.keys():
        lanly[key] = empty.copy()

    # get column indexes
    colID = pastDat[0].index('FormattedID')
    colPts = pastDat[0].index('PlanEstimate')
    colState = pastDat[0].index('ScheduleState')

    # process each row
    for datrow in pastDat[1:]:
        bDone = isDone(datrow[colState])
        # Process User Story
        if datrow[colID][0] == "U":
            lanly[datrow[colTeam]]['Stories.Total'] += 1
            if bDone:
                lanly[datrow[colTeam]]['Stories.Done'] += 1
            else:
                lanly[datrow[colTeam]]['Stories.NotDone'] += 1
            try:
                lanly[datrow[colTeam]]['Points.Total'] += datrow[colPts]
                if bDone: lanly[datrow[colTeam]]['Points.Done'] += datrow[colPts]
                else: lanly[datrow[colTeam]]['Points.NotDone'] += datrow[colPts]
            except TypeError:
                ptErrors += 1
                print("PtError ", ptErrors, datrow[colPts])
            if datrow[colPts] == 0:
                lanly[datrow[colTeam]]['Stories.Zero'] += 1
        # Process Defect
        elif datrow[colID][0] == "D":
            lanly[datrow[colTeam]]['Defects.Total'] += 1
            if bDone:
                lanly[datrow[colTeam]]['Defects.Done'] += 1
            else:
                lanly[datrow[colTeam]]['Defects.NotDone'] += 1

    # End Analysis
    for key in lanly.keys():
        if lanly[key]['Points.Total'] > 0:
            lanly[key]['Points.Percent'] = lanly[key]['Points.Done'] / lanly[key]['Points.Total']
        else:
            lanly[key]['Points.Percent'] = 0.0
        if lanly[key]['Stories.Total'] > 0:
            lanly[key]['Stories.Percent'] = lanly[key]['Stories.Done'] / lanly[key]['Stories.Total']
        else:
            lanly[key]['Stories.Percent'] = 0.0

    return lanly

def writeAnalysis(wBk, aDict):
    # create the sheet
    today = datetime.date.today()
    wsNameBase = ("Summary {:04d}.{:02d}.{:02d}").format(today.year, today.month, today.day)
    suffix = 'A'
    wsName = wsNameBase
    while wsName in wBk.sheetnames:
        wsName = wsNameBase + suffix
        suffix = chr(ord(suffix)+1)
    ws = wBk.create_sheet(wsName)

    ws.append( ["Project.Name"] + gAttribs )
    for key in aDict.keys():
        #ws.append([key] + [aDict[key][bkey] for bkey in aDict[key].keys()])
        # do it this way, not sure if above retains order
        ws.append([key] + [aDict[key][atr] for atr in gAttribs])

def Test09(wb):
    print ("Test 09 ", "-"*40)
    ws = wb.active #assume data on first sheet
    data09 = readWSData(ws)
    addExtendedWS(ws, data09)
    anly09 = analyzeData(data09)
    writeAnalysis(wb, anly09)

def Test08(wb):
    print ("Test 08 ", "-"*40)
    ws = wb.active #assume data on first sheet
    pp = pprint.PrettyPrinter(indent=4)
    data08 = readWSData(ws)

    anly08 = analyzeData(data08)
    writeAnalysis(wb, anly08)

def Test07(wb):
    # chr(ord('A')+1)
    today = datetime.date.today()
    wsNameBase = ("Summary {:04d}.{:02d}.{:02d}").format(today.year, today.month, today.day)
    suffix = 'A'
    wsName = wsNameBase
    print (wb.sheetnames)
    while wsName in wb.sheetnames:
        wsName = wsNameBase + suffix
        suffix = chr(ord(suffix)+1)
    ws = wb.create_sheet(wsName)
    print (wb.sheetnames)
    wb.save(fname)

def Test06():
    ptErrors = 0
    print ("Test 06 ", "-"*40)
    pp = pprint.PrettyPrinter(indent=4)

    colTeam = data[0].index("Project.Name")
    teamNames = list(set(row[colTeam] for row in data[1:]))

    empty = dict.fromkeys(gAttribs, 0)
    anly = dict.fromkeys(teamNames, None)
    for key in anly.keys():
        anly[key] = empty.copy()

    #for attrib in gAttribs:
    #    print (attrib, anly['Sabre Cruises'][attrib])
    colID = data[0].index('FormattedID')
    colPts = data[0].index('PlanEstimate')
    for datrow in data[1:]:
        if datrow[colID][0] == "U":
            anly[datrow[colTeam]]['Stories.Total'] += 1
            try:
                anly[datrow[colTeam]]['Points.Total'] += datrow[colPts]
            except TypeError:
                ptErrors += 1
                print("PtError ", ptErrors, datrow[colPts])
            if datrow[colPts] == 0:
                anly[datrow[colTeam]]['Stories.Zero'] += 1
        elif datrow[colID][0] == "D":
            anly[datrow[colTeam]]['Defects.Total'] += 1
    pp.pprint(anly)
    print("SPt errors = ", ptErrors)


def Test04():
    print ("Test 04 ", "-"*40)
    colTeam = data[0].index("Team")
    value = sum(row[0] for row in data[1:] if row[colTeam] == "Red" )
    print (value)

def Test05():
    print ("Test 05 ", "-"*40)
    #for row in data: print(row[0])
    colTeam = data[0].index("Team")
    teamNames = list(set(row[colTeam] for row in data[1:]))
    print (teamNames)
    value = dict.fromkeys(teamNames, 0)
    print (value)

    for row in data[1:]:
        value[row[colTeam]] += row[0]
    print (value)

def main():
    wb = load_workbook(fname)
    #ws = wb.active
    #
    #for row in ws.values:
    #    data.append(list(row))

    Test09(wb)
    #Test08(wb)
    #Test07(wb)
    #Test06()
    #Test05()
    #Test04()
    wb.save(fname)

if __name__ == '__main__':
    main()
    print ("Fini")
