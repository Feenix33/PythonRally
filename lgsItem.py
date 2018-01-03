from __future__ import print_function
import os
import sys
import csv
import getopt
import datetime
import argparse

from pyral import Rally, rallySettings, rallyWorkset
from openpyxl import Workbook

"""
---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ----
.Process FEA TF US
.Output the hierarchy
.User Story and defect processing
.Generic query
.Query from input file
.Input file with ignore
.Restructure output to have separate modules to write rather than as you go
.Add excel output
.what globals should be global
.should globals be in a structure
.excel page per FEA
.Consolidate query routines
.Commnand line switches

 Get tasks if asked for user stories and defects
 ini/yaml config file
 Better excel file creation - separation on workbooks by query token?
---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ----
 Error handling
   search not found
   files not opening
   incorrect ini
   unknown flags
   upper/lowercase
---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ----
.input file
.screen status or silent
.excel or csv
.output filename
.not found string
 tasks or not for all or only explicit us/de
---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ---- ----
"""


gFieldnames = ['FormattedID', 'Name', 'PlanEstimate', 'TeamFeature.FormattedID', 'TeamFeature.Name',
                        'ScheduleState', 'Project.Name', 'Iteration.Name', 'Owner.Name', 'CreationDate']
gPPMFieldnames = ['FormattedID', 'Name', 'LeafStoryPlanEstimateTotal', 'Parent.FormattedID', 'Parent.Name', 
                        'LeafStoryCount', 'AcceptedLeafStoryPlanEstimateTotal', 'AcceptedLeafStoryCount', 'Owner.Name', 'CreationDate']



class theConfig:
    def __init__(self):
        self.outToConsole = True
        self.writeCSV = True
        self.writeExcel = True
        self.fnameCSV = "lgsItem.csv"
        self.fnameExcel = "lgsItem.xlsx"
        self.fnameExceloutType = 0
        self.singleExcel = False
        self.fnameInput = 'in.txt'
        self.strNoExist = "--"
        self.useInput = True


gConfig = theConfig()


def convertIterIDtoLabel(iterID):
    # convert the short version of the iteration ID to the full version
    if len(iterID) != 3 and len(iterID) != 8:
        return iterID

    today = datetime.date.today()
    iterNum = int(iterID[1:3])
    if len(iterID) == 3: iterYear = today.year
    else: iterYear = int(iterID[4:])

    iter1start = datetime.date(2017, 1, 4)
    iter1end = datetime.date(2017, 1, 17)
    deltaDays = (iterYear - 2017) * 364 + (iterNum -1)*14
    iterStart = iter1start + datetime.timedelta(deltaDays)
    iterEnd = iterStart + datetime.timedelta(13)
    strIterLabel = ("S{:02d}#2017-{:02d}-{:02d}/{:02d}-{:02d}").format(
            iterNum, iterStart.month, iterStart.day, iterEnd.month, iterEnd.day)
    return strIterLabel


def processRecord(record):
    return  [getattrError(record, attr) for attr in gFieldnames] 

 
def getattrError(n, attr):
    try:
        return getattr(n, attr)
    except AttributeError:
        dot = attr.find(".")
        if dot == -1 :
            return str(gConfig.strNoExist)
        try:
            return getattr( getattr(n, attr[:dot]), attr[(dot+1):])
        except AttributeError:
            return str(gConfig.strNoExist)

def processIteration(instRally, nameIter, queryToken):
    records = []
    entityList = ['HierarchicalRequirement', 'Defect']
    #entityName = 'Defect'
    fullIterationName = convertIterIDtoLabel(nameIter[2:])
    queryString = 'Iteration.Name = "' + fullIterationName + '"'

    for entityName in entityList:
        response = instRally.get(entityName, fetch=True, projectScopeDown=True, query=queryString)
        #response = gRally.get(queryItem, fetch=True, projectScopeDown=True, query=gblIteration) 

        consoleStatus ("   " + queryString + "    Count = " + str(response.resultCount))
        for story in response:
            records.append( processRecord( story ) + [queryToken[2:]]) #xx
    return records

def processFEA(instRally, feaID, queryToken):
    entityName = 'PortfolioItem'
    queryString = 'FormattedID = "' + feaID + '"'
    response = instRally.get(entityName, fetch=True, projectScopeUp=True, query=queryString)
    records = []
    for ppmFeature in response:
        records.append( [getattrError(ppmFeature, attr) for attr in gPPMFieldnames] + [queryToken])
        if ppmFeature.DirectChildrenCount > 0 and hasattr(ppmFeature, 'Children'):
            for teamfeature in ppmFeature.Children:
                # add the team feature to the data table
                records.append( [getattrError(teamfeature, attr) for attr in gPPMFieldnames] + [queryToken])
                # add the subsequent children
                records.extend( processStory(instRally, teamfeature.FormattedID, queryToken))
    return records 

def processStory(instRally, storyID, queryToken): # story is user story or defect or TF
    records = []
    if storyID[0] == "U":
        entityName = 'HierarchicalRequirement'
        queryString = 'FormattedID = "' + storyID + '"'
    elif storyID[0] == "D":
        entityName = 'Defect'
        queryString = 'FormattedID = "' + storyID + '"'
    else:
        entityName = 'HierarchicalRequirement'
        queryString = 'Feature.FormattedID = "' + storyID + '"'
    response = instRally.get(entityName, fetch=True, projectScopeDown=True, query=queryString)
    consoleStatus ("   " + queryString + "    Count = " + str(response.resultCount))
    for story in response:
        records.append( processRecord( story ) + [queryToken]) #xx
    return records


def tokens(fileobj):
    for line in fileobj:
        if line[0] != '#':
            for word in line.split():
                yield word

def getHeadList(ppmType=None):
    #return the header items to write in csv or xlsx
    if ppmType: return (gPPMFieldnames + ['QueryToken'])
    else: return (gFieldnames + ['QueryToken'])

def writeCSV(listData, fname):
    fout = open(fname, 'wb')
    fwriter = csv.writer(fout, dialect='excel')
    fwriter.writerow( getHeadList() )
    fwriter.writerows(listData)
    fout.close()

def writeExcel(listData, fname, singleSheet):
    consoleStatus("Writing Excel")
    workbook = Workbook()
    ws = workbook.active
    queryTokenOld = "0None"
    queryToken    = listData[0][-1]

    if singleSheet:
        ws = workbook.create_sheet("Output")
        ws.append(gFieldnames + ['QueryToken'])

    for record in listData:
        queryToken = record[-1]
        if queryToken != queryTokenOld:
            if not singleSheet and (not ((queryToken[0] == 'U' or queryToken[0] == 'D') and (queryTokenOld[0] == 'U' or queryTokenOld[0] == 'D'))):
                sheetTitle = queryToken.replace('/','|') #excel no like slash
                #error need to check for duplicated sheet name
                ws = workbook.create_sheet(sheetTitle)
                queryTokenOld = queryToken
                if queryToken[0] == 'T' or queryToken[0] == 'F':
                    ws.append( getHeadList(True) )
                else:
                    ws.append( getHeadList() )
        ws.append(record)
    #get rid of the initial sheet
    std = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(std)
    workbook.save(fname)


def consoleStatus(message):
    if gConfig.outToConsole: print (message)

def buildInputParser(parser):
    parser.add_argument("-i", "--infile", type=argparse.FileType('r'), default='in.txt',
            help="input file with search parameters")
    parser.add_argument("-noi", "--noinput", default=False, action="store_true", help="No input file")
    parser.add_argument("-q", "--quiet", action="store_true", help="No console messages")
    parser.add_argument("-c", "--csv", nargs='?', const=1, type=str, default='out.csv', help="Write a .csv output")
    parser.add_argument("-x", "--xl", nargs='?', const=1, type=str, default='out.xlsx', help="Write a Excel output")
    parser.add_argument("--noxl", action='store_true', default=False, help="Suppress Excel output")
    parser.add_argument("--nocsv", action='store_true', default=False, help="Suppress csv output")
    parser.add_argument("-na", "--nastring", nargs=1, type=str, default="xx", help="String for empty")
    #parser.add_argument("--xltype", nargs='?', const=1, type=int, default=1, help="Excel file output format")
    parser.add_argument("-xl1", "--xlone", action='store_true', default=False, help="Excel file output as single sheet")
    parser.add_argument('baz',default=[], nargs='*', help="Parameters to search for")

def mapArgsserToGlobal(args):
    #convert the passed arguments to the global configuration and return the search tokens
    gConfig.outToConsole = not args.quiet
    gConfig.fnameInput = args.infile.name
    gConfig.writeExcel = True
    gConfig.writeCSV = True

    if args.xl == 1: gConfig.fnameExcel = "lgsItem.xlsx"
    else: 
        gConfig.fnameExcel = args.xl
        if gConfig.fnameExcel[-5:] != '.xlsx': gConfig.fnameExcel += '.xlsx'


    if args.csv == 1: gConfig.fnameCSV = "lgsItem.csv"
    else: gConfig.fnameCSV = args.csv

    if args.noxl: gConfig.writeExcel = False
    if args.nocsv: gConfig.writeCSV = False

    gConfig.strNoExist = args.nastring[0]
    #gConfig.fnameExceloutType = args.xltype
    gConfig.useInput = not args.noinput
    #print (args)
    return args.baz


def main():
    dataHR = []
    queryList = []

    inParser = argparse.ArgumentParser()
    buildInputParser(inParser)
    args = inParser.parse_args()
    queryList.extend(mapArgsserToGlobal(args))

    finputfile = gConfig.fnameInput
    if gConfig.useInput:
        consoleStatus('Getting search tokens...')
        try:
            infile = open(finputfile, 'r')
            queryList.extend(tokens(infile))

        except IOError:
            print ("IOError: Cannot open", sys.argv[0], "<input file>")


    consoleStatus('Logging in...')
    rally = Rally(server, apikey=apikey, workspace=workspace, project=project)

    consoleStatus('Query execution...')

    for queryItem in queryList:
        if queryItem[:2] == "FE":
            dataHR.extend (processFEA(rally, queryItem, queryItem))
        elif queryItem[:2] == "TF":
            dataHR.extend (processStory(rally, queryItem, queryItem))
        elif queryItem[:2] == "US":
            dataHR.extend (processStory(rally, queryItem, queryItem))
        elif queryItem[:2] == "DE":
            dataHR.extend (processStory(rally, queryItem, queryItem))
        elif queryItem[:2] == "IT":
            dataHR.extend (processIteration(rally, queryItem, queryItem))
        else:
            print ("Error query for " + queryItem)

    if gConfig.writeCSV: writeCSV(dataHR, gConfig.fnameCSV)
    if gConfig.writeExcel: writeExcel(dataHR, gConfig.fnameExcel, gConfig.singleExcel)
    consoleStatus('Fini')



if __name__ == '__main__':
    execfile("apifig.py")
    main()
