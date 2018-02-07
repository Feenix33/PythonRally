"""
ppIteration - Post Process a Rally Iteration dump

Adds another sheet summarizing the iteration data and adds columns to the iteration data
to aid in processing: 
    Type - first letters of the Formatted ID
    Done - flag if done or not
    Iteration Name - convert iteration name to a sortable name

----------- ----------- ----------- -----------
Todo:
    . Need for py2?
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import pprint
import datetime
import argparse

gAttribs = [
        'Stories.Total', 'Stories.Zero', 'Stories.Done', 'Stories.NotDone', 'Stories.Percent',
        'Defects.Total', 'Defects.Done', 'Defects.NotDone',
        'Points.Total', 'Points.Done', 'Points.NotDone', 'Points.Percent'
    ]

############
#gFieldnames = ['FormattedID', 'Name', 'PlanEstimate', 'TeamFeature.FormattedID', 
#        'TeamFeature.Name', 'ScheduleState', 'Project.Name', 'Iteration.Name', 
#        'Owner.Name', 'CreationDate']
#gDoneStates = ['Accepted','Completed','Released-to-Production']

def findWorksheet(wb):
    #find the data worksheet, use FormmatedID in first cell as detector
    #assume first sheet then search if not there
    ws = wb.active #assume data on first sheet
    if ws.cell(row=1,column=1).value == 'FormattedID':
        return ws
    #did not find so loop through all sheets and try to find
    for wsName in wb.sheetnames:
        ws = wb[wsName]
        if ws.cell(row=1,column=1).value == 'FormattedID':
            return ws
    return None

def addExtendedWS(ws, inDat):
    #use exception handling to punch out if already done
    try:
        noNeed = inDat[0].index('Type')
        return
    except:
        pass
    colID = inDat[0].index('FormattedID')
    colState = inDat[0].index('ScheduleState')
    colIter = inDat[0].index('Iteration.Name')

    outType = ws.max_column + 1
    outDone = outType + 1
    outIter = outDone + 1
    atRow = 2
    ws.cell(row=1, column=outType).value = "Type"
    ws.cell(row=1, column=outDone).value = "Done"
    ws.cell(row=1, column=outIter).value = "Iteration.Sortable"

    for row in inDat[1:]: # have to write directly to sheet
        ws.cell(row=atRow, column=outType).value = row[colID][0:2]
        ws.cell(row=atRow, column=outDone).value = isDone(row[colState])
        try:
            ws.cell(row=atRow, column=outIter).value = row[colIter][4:8]+"#"+row[colIter][0:3]
        except TypeError:
            pass
        atRow += 1

def isDone(state):
    #lDoneStates = ['Accepted','Completed','Released-to-Production']
    #return state in lDoneStates
    return state in ['Accepted','Completed','Released-to-Production']

def readWSData(ws):
    localData = []
    for row in ws.values:
        localData.append(list(row))
    return localData

def analyzeData(pastDat):
    colTeam = pastDat[0].index("Project.Name")
    teamNames = list(set(row[colTeam] for row in pastDat[1:]))

    empty = dict.fromkeys(gAttribs, 0)
    localAnalysis = dict.fromkeys(teamNames, None)
    for key in localAnalysis.keys():
        localAnalysis[key] = empty.copy()

    # get column indexes
    colID = pastDat[0].index('FormattedID')
    colPts = pastDat[0].index('PlanEstimate')
    colState = pastDat[0].index('ScheduleState')

    # process each row
    for datrow in pastDat[1:]:
        bDone = isDone(datrow[colState])
        # Process User Story
        if datrow[colID][0] == "U":
            localAnalysis[datrow[colTeam]]['Stories.Total'] += 1
            if bDone:
                localAnalysis[datrow[colTeam]]['Stories.Done'] += 1
            else:
                localAnalysis[datrow[colTeam]]['Stories.NotDone'] += 1
            try:
                localAnalysis[datrow[colTeam]]['Points.Total'] += datrow[colPts]
                if bDone: localAnalysis[datrow[colTeam]]['Points.Done'] += datrow[colPts]
                else: localAnalysis[datrow[colTeam]]['Points.NotDone'] += datrow[colPts]
            except TypeError:
                ptErrors += 1
                print("PtError ", ptErrors, datrow[colPts])
            if datrow[colPts] == 0:
                localAnalysis[datrow[colTeam]]['Stories.Zero'] += 1
        # Process Defect
        elif datrow[colID][0] == "D":
            localAnalysis[datrow[colTeam]]['Defects.Total'] += 1
            if bDone:
                localAnalysis[datrow[colTeam]]['Defects.Done'] += 1
            else:
                localAnalysis[datrow[colTeam]]['Defects.NotDone'] += 1

    # End Analysis
    for key in localAnalysis.keys():
        if localAnalysis[key]['Points.Total'] > 0:
            localAnalysis[key]['Points.Percent'] = localAnalysis[key]['Points.Done'] / localAnalysis[key]['Points.Total']
        else:
            localAnalysis[key]['Points.Percent'] = 0.0
        if localAnalysis[key]['Stories.Total'] > 0:
            localAnalysis[key]['Stories.Percent'] = localAnalysis[key]['Stories.Done'] / localAnalysis[key]['Stories.Total']
        else:
            localAnalysis[key]['Stories.Percent'] = 0.0

    return localAnalysis

def writeAnalysis(wBk, aDict):
    # create the sheet
    today = datetime.date.today()
    wsNameBase = ("Summary {:04d}.{:02d}.{:02d}").format(today.year, today.month, today.day)
    suffix = 'A'
    wsName = wsNameBase
    while wsName in wBk.sheetnames:
        wsName = wsNameBase + suffix
        suffix = chr(ord(suffix)+1)
    ws = wBk.create_sheet(wsName)

    ws.append( ["Project.Name"] + gAttribs )
    for key in aDict.keys():
        #ws.append([key] + [aDict[key][bkey] for bkey in aDict[key].keys()])
        # do it this way, not sure if above retains order
        ws.append([key] + [aDict[key][atr] for atr in gAttribs])

def buildInputParser(parser):
    parser.add_argument('-a', "--addit",default=False, action='store_true', 
            help="Add additional processing columns")
    parser.add_argument('-ns', "--nosum",default=False, action='store_true', 
            help="Do not do summary processing")
    parser.add_argument('inFileName', help="Excel file to process")

def main():
    inParser = argparse.ArgumentParser(description="Post Process an Iteration")
    buildInputParser(inParser)
    args = inParser.parse_args()

    localFileName = args.inFileName
    lWorkbook = load_workbook(localFileName)

    #lWorksheet = lWorkbook.active #assume data on first sheet
    lWorksheet = findWorksheet(lWorkbook)
    if lWorksheet == None:
        print("Error: Cannot find data worksheet in file")
        quit()


    data09 = readWSData(lWorksheet)
    if args.addit: addExtendedWS(lWorksheet, data09)
    if not args.nosum:
        anly09 = analyzeData(data09)
        writeAnalysis(lWorkbook, anly09)

    lWorkbook.save(localFileName)


if __name__ == '__main__':
    main()
